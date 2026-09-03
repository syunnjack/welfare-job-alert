"""総務省の「全国地方公共団体コード」から、自治体マスタを作る。

出典: https://www.soumu.go.jp/denshijiti/code.html
      都道府県コード及び市区町村コード（Excel）

このマスタが、自治体ごとの障害福祉情報を集める土台になる。
**名称とコードは公式のものをそのまま使う。**推測で補わない。
政令指定都市の行政区は別シートにあるが、制度は市単位で決まるため
ここでは市までを対象にする。

使い方:
  python scripts/fetch-municipalities.py
"""
import json
import sys
import urllib.request
from pathlib import Path

import openpyxl

SOURCE_URL = 'https://www.soumu.go.jp/main_content/000925835.xlsx'
SOURCE_PAGE = 'https://www.soumu.go.jp/denshijiti/code.html'
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'municipal-codes.xlsx'
OUT = ROOT / 'data' / 'municipalities.json'


def download() -> None:
    request = urllib.request.Request(SOURCE_URL, headers={'User-Agent': 'welfarejob.jp municipality fetcher'})
    with urllib.request.urlopen(request, timeout=60) as response:
        XLSX.write_bytes(response.read())
    print(f'取得しました: {XLSX.name} ({XLSX.stat().st_size:,}バイト)')


def main() -> None:
    if '--offline' not in sys.argv:
        download()

    workbook = openpyxl.load_workbook(XLSX)
    sheet = workbook[workbook.sheetnames[0]]

    prefectures = []
    municipalities = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, pref, city, pref_kana, city_kana = (row + (None,) * 5)[:5]
        if not code or not pref:
            continue
        code = str(code).strip()
        pref = str(pref).strip()

        if not city:
            # 市区町村名が空の行は都道府県そのもの。
            prefectures.append({'code': code, 'name': pref, 'kana': (pref_kana or '').strip()})
            continue

        municipalities.append({
            'code': code,
            'pref': pref,
            'name': str(city).strip(),
            'kana': (city_kana or '').strip(),
        })

    if len(prefectures) != 47:
        print(f'都道府県が47件になりません（{len(prefectures)}件）。取り込みを中止します。', file=sys.stderr)
        raise SystemExit(1)

    data = {
        'source': SOURCE_PAGE,
        'sourceFile': SOURCE_URL,
        'prefectures': prefectures,
        'municipalities': municipalities,
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=1) + '\n', encoding='utf-8')
    print(f'都道府県 {len(prefectures)}件 / 市区町村 {len(municipalities)}件 を {OUT.name} に書き出しました。')


if __name__ == '__main__':
    main()
